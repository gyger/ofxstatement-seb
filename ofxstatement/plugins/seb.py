# -*- coding: utf-8 -*-

import re
import itertools
import logging
import locale

from datetime import datetime
from openpyxl import load_workbook

from contextlib import contextmanager
from ofxstatement.parser import StatementParser
from ofxstatement.plugin import Plugin
from ofxstatement.statement import Statement, StatementLine, generate_transaction_id


def take(n, iterable):
    """Return first n items of the iterable as a list."""
    return list(itertools.islice(iterable, n))


@contextmanager
def scoped_setlocale(category, loc=None):
    """Scoped version of locale.setlocale()"""
    orig = locale.getlocale(category)
    try:
        yield locale.setlocale(category, loc)
    finally:
        locale.setlocale(category, orig)


def atof(string, loc=None):
    """Locale aware atof function for our parser."""
    with scoped_setlocale(locale.LC_NUMERIC, loc):
        return locale.atof(string)


class SebStatementParser(StatementParser):
    date_format = '%Y-%m-%d'
    bank_id = 'SEB'
    currency_id = 'SEK'
    footer_regexps = [
        '^Datum:  -',
        '^Datum: ([0-9]{4}-[0-9]{2}-[0-9]{2}) - ([0-9]{4}-[0-9]{2}-[0-9]{2})$'
    ]

    def __init__(self, fin, locale=None, brief=False):
        """
        Create a new SebStatementParser instance.

        :param fin: filename to create parser for
        :param brief: whenever to attempt replace description with a brief version i.e. all extra info removed
        """

        self.locale = locale
        self.brief = brief

        self.workbook = load_workbook(filename=fin, read_only=True)
        self.validate()

        self.statement = self.parse_statement()

    def validate(self):
        """
        Naive validation to make sure that xlsx document is structured the way it was
        when this parser was written.

        :raises ValueError if workbook has invalid format
        """

        try:
            self._validate()
        except AssertionError as e:
            raise ValueError(e)

    def _validate(self):
        sheet = self.workbook.active

        logging.info('Checking that sheet has at least 8 rows.')
        rows = take(8, sheet.iter_rows())
        assert len(rows) == 8

        logging.info('Verifying that every row has 6 cells.')
        assert type(rows) == list
        for row in rows:
            assert len(row) == 6

        logging.info('Extracting values from every cell.')
        rows = [[c.value for c in row] for row in rows]

        logging.info('Verifying account id.')
        summary_account_row = rows[4]
        account_id = summary_account_row[0]
        assert re.match('^\w+\s\(([0-9\s]+)\)$', account_id)
        assert [None, None] == summary_account_row[-2:]

        def is_footer(row):
            for r in self.footer_regexps:
                if re.match(r, row[0]):
                    return True
            return False

        logging.info('Verifying empty row.')
        empty_row = rows[5]
        assert [None, None, None, None, None, None] == empty_row

        logging.info('Verifying statements header.')
        statement_header_row = rows[7]
        assert ['Bokföringsdatum', 'Valutadatum', 'Verifikationsnummer', 'Text', 'Belopp', 'Saldo'] == statement_header_row

        logging.info('Everything is OK!')

    def parse_statement(self):
        statement = Statement()
        sheet = self.workbook.active

        # We need first 8 rows
        rows = take(8, sheet.iter_rows())
        rows = [[c.value for c in row] for row in rows]

        assert len(rows) == 8
        account_row = rows[4]

        account_id = re.match('^\w+\s\(([0-9\s]+)\)$', account_row[0])
        statement.account_id = account_id[1]
        statement.bank_id = self.bank_id
        statement.currency = self.currency_id
        
        return statement

    def split_records(self):
        sheet = self.workbook.active

        # Skip first 8 rows. Headers they are.
        for row in itertools.islice(sheet.iter_rows(), 8, None):
            yield [c.value for c in row]

    def parse_record(self, row):
        row = take(5, row)

        stmt_line = StatementLine()
        stmt_line.date = self.parse_datetime(row[0])
        _ = self.parse_datetime(row[1])  # TODO: ???
        stmt_line.refnum = row[2]
        stmt_line.memo = row[3]
        stmt_line.amount = row[4]

        #
        # Looks like SEB formats description for card transactions so it includes the actual purchase date
        # within e.g. 'WIRSTRÖMS PU/14-12-31' and it means that description is 'WIRSTRÖMS PU' while the actual
        # card operation is 2014-12-31.
        #
        # P.S. Wirströms Irish Pub is our favorite pub in Stockholm: http://www.wirstromspub.se
        #
        m = re.match('(.*)/([0-9]{2}-[0-9]{2}-[0-9]{2})$', stmt_line.memo)
        if m:
            card_memo, card_date = m.groups()
            if self.brief:
                stmt_line.memo = card_memo
            stmt_line.date_user = datetime.strptime(card_date, '%y-%m-%d')

        stmt_line.id = generate_transaction_id(stmt_line)
        return stmt_line


def parse_bool(value):
    if value in ('True', 'true', '1'):
        return True
    if value in ('False', 'false', '0'):
        return False
    raise ValueError("Can't parse boolean value: %s" % value)


class SebPlugin(Plugin):
    def get_parser(self, fin):
        kwargs = {
            'locale': 'sv_SE'
        }
        if self.settings:
            if 'locale' in self.settings:
                kwargs['locale'] = parse_bool(self.settings.get('locale'))
            if 'brief' in self.settings:
                kwargs['brief'] = parse_bool(self.settings.get('brief'))
        return SebStatementParser(fin, **kwargs)
